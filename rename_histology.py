#!/usr/bin/env python3

"""
Copy and re-organize histology data in a new directory, whose structure
is consistent with the filesystem hierarchy required by cloud storage.

This script is based on:
https://github.com/faryabiLab/hpap-apps/blob/master/data_curator_tools/psv_pipelines/psv_histology_upload.py

The original version uses Pandas to parse the Excel file and match image
files. It was an overkill and made the code harder to understand, so
`pandas` was replaced by `openpyxl` when Excel file was read in.
"""

import os
import re
import shutil
import sys
import time

import openpyxl

IMG_FILE_EXTENSION = '.ndpi' # extension of image files

def my_log(message, with_time=True):
    """Print a log message, with a prefix of current time."""

    if with_time:
        message = f"{time.ctime()}: {message}"

    print(message, flush=True)


def get_filename_key(input_filename, rm_extension):
    """
    Return a filename key, which includes only '_' and alphanumric in
    `input_filename`. If `rm_extension` is True, the extension will be
    also removed from the key before it's returned.
    """

    if rm_extension:
       input_filename = input_filename.rsplit('.', 1)[0]

    filename_key = ""
    for c in input_filename:
        if c == '_' or c == '-' or c.isalpha() or c.isdigit():
            filename_key += c

    if len(filename_key) == 0 or not filename_key.startswith('HPAP'):
        my_log("ERROR: invalid filename '{input_filename}'")
        sys.exit(1)

    return filename_key


def get_donor_id(input_str):
    """
    Return the donor ID from input_str.
    (assume thqat `input_str` is in the format of "HPAP xxx_whatever".
    """

    hpap_str = input_str.split('_', 1)[0]
    num_part = hpap_str.split('HPAP')[1]
    donor_id = num_part.zfill(3)

    return donor_id


def check_img_filenames(filenames):
    """
    Check input image filenames.
    If any of them has different HPAP donor ID, print an error and exit
    immediately.
    If everything is good, return the donor_id and a dict (whose key is
    normalized filename key, value is the actual image filename).
    """

    donor_id = None
    fn_map = dict()
    for fn in filenames:
        fn_key = get_filename_key(fn, rm_extension=True)
        current_donor_id = get_donor_id(fn_key)
        if donor_id is None:
            donor_id = current_donor_id
        elif donor_id != current_donor_id:
            print(f"ERROR: inconsistent donor ID in '{fn}'")
            sys.exit(1)

        if fn_key in fn_map:
            print(f"ERROR: duplicate filename key in '{fn}'")
            sys.exit(1)

        fn_map[fn_key] = fn

    return donor_id, fn_map


def check_src(src_dir):
    """Ensure that source data directory is in good shape."""

    src_dir = src_dir.replace("\\", "/").replace('"', '')
    if src_dir.endswith('/'):  # remove trailing '/'
        src_dir = src_dir[:-1]

    img_files = list()
    excel_files = list()
    for x in os.listdir(src_dir):
        if x.endswith(IMG_FILE_EXTENSION):
            img_files.append(x)
        elif x.endswith('.xlsx'):
            excel_files.append(x)

    # Make sure that one and only one Excel file is found in `src_dir`:
    if len(excel_files) == 0:
        my_log(f"ERROR: Excel file not found in {src_dir}")
        sys.exit(1)

    if len(excel_files) != 1:
        my_log(f"ERROR: multiple Excel files found in {src_dir}")
        sys.exit(1)

    num_img = len(img_files)
    if num_img == 0:
        my_log(f"ERROR: no image file found in '{src_dir}'")
        sys.exit(1)

    # Ensure that image filenames have identical donor IDs
    donor_id, fn_map = check_img_filenames(img_files)

    print(f"{num_img} image file(s) found in '{src_dir}'")
    excel_filename = f"{src_dir}/{excel_files[0]}"

    return donor_id, fn_map, excel_filename


def check_dest(dest_dir):
    """
    Ensure that destination directory either does not exist, or
    is an empty directory.
    """

    if not os.path.exists(dest_dir):
        return

    if not os.path.isdir(dest_dir):
        print(f"ERROR: '{dest_dir}' exists but is not a directory")
        sys.exit(3)

    if len(os.listdir(dest_dir)):
        print(f"ERROR: '{dest_dir}' is not an empty directory")
        sys.exit(4)


def check_excel_filenames(rows, excel_filename):
    """
    Ensure that each row in Excel includes same donor ID. If it does,
    return this donor ID.
    """

    donor_id = None
    for r, v in rows.items():
        current_donor_id = get_donor_id(r)
        if donor_id is None:
            donor_id = current_donor_id
        elif donor_id != current_donor_id:
            excel_val = v['filename_stem']
            print("ERROR: inconsistent donor ID in '{excel_val}' of '{excel_filename}'")
            sys.exit(1)

    return donor_id


def get_excel_columns(sheet_obj):
    """
    Read the first data row (not the one with column names) of input
    Excel sheet object, and return a pair of column numbers (integers),
    the first is HPAP column's number in Excel file, the second is
    'Prep' column's number in Excel file.
    """

    max_col = sheet_obj.max_column

    hpap_col = prep_col = None
    for c in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=2, column=c)
        cell_val = cell_obj.value

        # If a column's value is empty or not a string, skip it.
        if cell_val is None or not isinstance(cell_val, str):
            continue

        cell_val = cell_val.strip()
        lower_val = cell_val.lower()

        # Skip columns that includes whitespace characters only
        if len(lower_val) == 0:
            continue

        if cell_val.startswith('HPAP'):
            if hpap_col is None:
                hpap_col = c
                continue
            else:
                print(
                    f"ERROR in Excel file: columns #{hpap_col} and #{c} of row "
                    f"#2 both look like 'HPAP' column"
                )
                sys.exit(1)

        if 'oct' in lower_val or 'ffpe' in lower_val or 'vand' in lower_val:
            if prep_col is None:
                prep_col = c
                continue
            else:
                print(
                    f"ERROR in Excel file: columns #{hpap_col} and #{c} of row "
                    f"#2 both look like 'Prep' column"
                )
                sys.exit(1)

    if hpap_col is None:
        print("ERROR in Excel file: 'HPAP' column not found")
        sys.exit(1)

    if hpap_col is None:
        print("ERROR in Excel file: 'Prep' column not found")
        sys.exit(1)

    return hpap_col, prep_col


def rename_stain(input_str):
    input_str = input_str.strip().upper()

    if input_str == "OCT":
        return 'OCT-flash-frozen'

    if 'VAN' in input_str:
        return 'OCT-lightly-fixed'

    return input_str

def read_excel(filename, donor_id):
    """
    Read Excel spreadsheet, parse the data in each cell, and return a
    dict, whose key is a "normalized" filename key, and whose value is
    another dict of other columns.
    """

    # workbook and sheet objects
    wb_obj = openpyxl.load_workbook(filename, read_only=True)
    sheet_obj = wb_obj.active
    hpap_col, prep_col = get_excel_columns(sheet_obj)

    # Read the whole Excel file and create a dict, whose key is the value
    # of `filename_stem`, and whose value is another dict of other columns.
    max_row = sheet_obj.max_row
    rows = dict()
    for r in range(2, max_row + 1):
        curr_row = dict()
        hpap_col_val = sheet_obj.cell(row=r, column=hpap_col).value.strip()
        row_key = get_filename_key(hpap_col_val, rm_extension=False)
        if len(row_key) == 0 or not row_key.startswith('HPAP'):
            print(f"ERROR: invalid HPAP value on row #{r} column #{hpap_col}")
            sys.exit(1)

        if row_key in rows:
            print(f"ERROR: duplicate HPAP value on row #{r} of '{filename_stem}'")
            sys.exit(1)

        curr_row['filename_stem'] = hpap_col_val

        prep_col_val = sheet_obj.cell(row=r, column=prep_col).value.strip()
        curr_row['stain'] = rename_stain(prep_col_val)
        rows[row_key] = curr_row

    excel_donor_id = check_excel_filenames(rows, filename)
    if excel_donor_id != donor_id:
        print(
            f"ERROR: donor ID in Excel ({excel_donor_id}) does not match "
            f"the one in image files ({donor_id})"
        )
        sys.exit(1)

    return rows


def search_pancreas(input_str):
    re_matches = re.findall("(pancreas)\\s?-?\\s?(\\w+)", input_str)
    if not re_matches:
        return

    re_matches = re_matches[0]
    if len(re_matches) == 1:
        return 'Pancreas', 'Pancreas'

    if len(re_matches) == 2:
        if re_matches[1] == 'unsure':
            return 'Pancreas', 'Pancreas-Unsure-of-orientation',

        if re_matches[1] in ['head', 'body', 'tail']:
            cap_str = re_matches[1].capitalize()
            return 'Pancreas', f'{cap_str}-of-pancreas'


def search_duodenum(input_str, needle):
    re_matches = re.findall(f"({needle})\\s?-?\\s?(\\w+)", input_str)
    if not re_matches:
        return

    re_matches = re_matches[0]
    if len(re_matches) == 1:
        return 'Duodenum', 'Duodenum'

    if len(re_matches) == 2:
        if re_matches[1] == 'unsure':
            return 'Duodenum', 'Duodenum-Unsure-of-orientation'

        if re_matches[1] in ['distal', 'mid', 'proximal']:
            cap_str = re_matches[1].capitalize()
            return 'Duodenum', f'Duodenum-{cap_str}-one-third'

        if re_matches[1] == 'prox':
            return 'Duodenum', 'Duodenum-Proximal-one-third'


def search_lymph_node(input_str):
    re_matches = re.findall("(ln)\\s?-?\\s?(\\w+)?", input_str)
    if not re_matches:
        return

    re_matches = re_matches[0]
    if len(re_matches) == 1:
        return 'Lymph node', 'Lymph-node'

    if len(re_matches) == 2:
        if re_matches[1] == 'sma':
            return 'Lymph node', 'Lymph-node-SMA'

        if re_matches[1] in ['body', 'head', 'tail']:
            cap_str = re_matches[1].capitalize()
            return 'Lymph node', f'Lymph-node-{cap_str}-of-pancreas'

        if re_matches[1] in ['mesentery', 'mesentary', 'mestentery']:
            return 'Lymph node', 'Lymph-node-Mesentery'


def get_anatomy_names(input_str):
    """
    Return a pair of strings, the first is short anatomy name (which will
    be the parent directory's name), the second is long anatomy name
    (which will be used in the new image file's name).
    """

    lower_str = input_str.lower()

    if 'spleen' in lower_str:
        return 'Spleen', 'Spleen'

    if 'thymus' in lower_str:
        return 'Thymus', 'Thymus'

    if 'artery' in lower_str:
        return 'Artery', 'Artery'

    # Search 'pancreas'
    search_result = search_pancreas(lower_str)
    if search_result:
        return search_result

    # Search 'duodenum'
    search_result = search_duodenum(lower_str, 'duodenum')
    if search_result:
        return search_result

    # Search "duod"
    search_result = search_duodenum(lower_str, 'duod')
    if search_result:
        return search_result

    # Search 'ln' (lymph node)
    search_result = search_lymph_node(lower_str)
    if search_result:
        return search_result

    print(f"ERROR: valid anatomy name not found in '{input_str}'")
    sys.exit(1)


def map_excel_to_images(excel_dict, img_dict):
    """
    Match the data in Excel with actual image files.
    If everything is good, return a another dict maps each image file's
    path to a destination file's name.
    """

    # Ensure that each image file matches one row in Excel spreadsheet.
    for ik, iv in img_dict.items():
        if ik not in excel_dict:
            print(f"ERROR: image file '{iv}' not match any row in Excel")
            sys.exit(1)

    # Ensure that each row in Excel matches one image file
    src2dest = dict()
    dest_key_counter = dict()
    for xk, xv in excel_dict.items():
        cell_v = xv['filename_stem']
        if xk not in img_dict:
            print(f"ERROR: '{cell_v}' not match any image files")
            sys.exit(1)

        short_anatomy, long_anatomy = get_anatomy_names(cell_v)
        stain = xv['stain']
        dest_key = "_".join(
            [
                f"HPAP-{donor_id}",
                "Histology",
                long_anatomy,
                stain,
                "H-and-E",
            ]
        )

        counter = dest_key_counter.get(dest_key, 0)
        uniq_num = counter + 1
        dest_key_counter[dest_key] = uniq_num

        dest_name = f"{dest_key}_{uniq_num}{IMG_FILE_EXTENSION}"

        src_name = img_dict[xk]
        src2dest[src_name] = {
            'sub_dir': short_anatomy,
            'name': dest_name,
        }

    return src2dest


def copy_src_to_dest(src_dir, dest_dir, donor_id, src2dest):
    """Copy source image files to the destination directory."""

    os.makedirs(dest_dir, exist_ok=True)

    for src_name, dest_dict in src2dest.items():
        src_path = os.path.join(src_dir, src_name)
        sub_dir = dest_dict['sub_dir']
        dest_parent = os.path.join(
            dest_dir, f'HPAP-{donor_id}', 'Histology', sub_dir
        )

        os.makedirs(dest_parent, exist_ok=True)
        dest_name = dest_dict['name']
        dest_path = os.path.join(dest_parent, dest_name)

        my_log(f"Copying '{src_name}' ...")
        shutil.copy(src_path, dest_path)

# ============================ Main program ==================================

if __name__ == "__main__":
    args = sys.argv

    if len(args) != 3:
        print("Usage: rename_histology.py <source_data_dir> <target_data_dir>")
        sys.exit(1)

    # Parse arguments
    src_dir = args[1]
    dest_dir = args[2]

    # Make sure that source directory is good
    donor_id, img_filenames, excel_filename = check_src(src_dir)

    excel_rows = read_excel(excel_filename, donor_id)

    #import json; print(json.dumps(rows, indent=2)); sys.exit(0)  # dhu test

    # Make sure that destination directory is good:
    check_dest(dest_dir)

    # Create a map between source image file and destination image file
    src2dest = map_excel_to_images(excel_rows, img_filenames)

    # Copy image files from source to destination
    copy_src_to_dest(src_dir, dest_dir, donor_id, src2dest)

    my_log("Done!")
